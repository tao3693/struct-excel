from dataclasses import fields, is_dataclass
from enum import Enum
from struct_excel.normalization import normalize_sheet
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from struct_excel.reader import read_raw_row
from struct_excel.transform import (
    to_course,
    to_enrollment,
    to_session,
    to_student,
    to_supervisor,
)
import logging
import argparse
from pathlib import Path


def dataclass_to_row(obj):
    row = []
    for f in fields(obj):
        value = getattr(obj, f.name)
        if isinstance(value, Enum):
            value = value.value
        row.append(value)

    return row


def write_dataclass_sheet(workbook: Workbook, sheet_name: str, data: list):
    ws = workbook.create_sheet(title=sheet_name)

    if not data:
        return

    if not is_dataclass(data[0]):
        raise ValueError(f"{sheet_name} contains non-dataclass objects")

    # Title
    headers = [f.name for f in fields(data[0])]
    ws.append(headers)

    # Data
    for obj in data:
        ws.append(dataclass_to_row(obj))


def parse_args():
    parser = argparse.ArgumentParser(
        description="Process Excel workbook into normalized outputs."
    )
    parser.add_argument(
        "--src",
        required=True,
        help="Path to source Excel file",
    )

    return parser.parse_args()


def main():
    dist_dir = Path("dist")
    dist_dir.mkdir(parents=True, exist_ok=True)

    ERR_XLSX = "./dist/err.xlsx"

    args = parse_args()

    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    logger = logging.getLogger(__name__)
    wb = load_workbook(args.src)
    err_wb = Workbook()

    for sheet in wb:
        err_ws = err_wb.create_sheet(sheet.title)
        normalize_sheet(sheet, err_ws)

    err_wb.save(ERR_XLSX)

    for sheet in wb:
        output = Workbook()
        output.remove(output.active)  # pyright: ignore
        output_dir = f"dist/{sheet.title}_output.xlsx"

        try:
            raw_rows = read_raw_row(sheet)
        except ValueError as e:
            logger.error(str(e))
            raise

        supervisors = to_supervisor(raw_rows)
        courses = to_course(raw_rows)
        students = to_student(raw_rows, supervisors)
        sessions = to_session(raw_rows, courses)
        enrollments = to_enrollment(raw_rows, students, courses, sessions)

        write_dataclass_sheet(output, "supervisors", supervisors)
        write_dataclass_sheet(output, "courses", courses)
        write_dataclass_sheet(output, "students", students)
        write_dataclass_sheet(output, "sessions", sessions)
        write_dataclass_sheet(output, "enrollments", enrollments)
        output.save(output_dir)
        print(f"Saved to {output_dir}, and errors saved to {ERR_XLSX}")


if __name__ == "__main__":
    main()
