import logging
import re
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import pycountry
import phonenumbers

logger = logging.getLogger(__name__)

_ALIASES: dict[str, str] = {
    "uk": "GB",
    "u.k.": "GB",
    "england": "GB",
    "scotland": "GB",
    "usa": "US",
    "u.s.": "US",
    "america": "US",
}

_INVALID_COUNTRY = "UNKNOWN"
_INVALID_PHONE = "NAN"


def _get_row_data(ws: Worksheet, row: int) -> list:
    return [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]


def _add_error_row(
    ws: Worksheet,
    err_ws: Worksheet,
    row: int,
    rows_to_delete: list,
    reason: str,
) -> None:
    row_data = _get_row_data(ws, row)
    row_data.append(reason)
    err_ws.append(row_data)
    rows_to_delete.append(row)


def normalize_sheet(ws: Worksheet, err_ws: Worksheet) -> None:
    headers = [cell.value for cell in ws[1]]
    headers.append("Error Reason")
    ws = _trim_cells(ws)
    err_ws.append(headers)

    try:
        _normalize_country_and_phone_col(ws, err_ws)
    except Exception:
        pass


def _trim_cells(ws: Worksheet) -> Worksheet:
    for row in ws:
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                value = re.sub(r"[\u200b\u200c\u200d\ufeff\u00ad]", "", cell.value)
                cell.value = value.strip()

    return ws


def _normalize_country_and_phone_col(ws: Worksheet, err_ws: Worksheet) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}

    country_col = headers.get("Country")
    if not country_col:
        raise ValueError("Country column not found")

    phone_col = headers.get("Phone")
    if not phone_col:
        raise ValueError("Phone column not found")

    rows_to_delete = []

    for row in range(2, ws.max_row + 1):
        country_cell = ws.cell(row=row, column=country_col)
        phone_cell = ws.cell(row=row, column=phone_col)
        normalized_country = ""

        if not isinstance(country_cell, Cell) or not isinstance(phone_cell, Cell):
            continue

        normalized_country = _normalize_country(country_cell.value)
        if normalized_country == _INVALID_COUNTRY:
            logger.error(
                f"row: {row} exception: normalize country={country_cell.value} failed"
            )
            _add_error_row(
                ws,
                err_ws,
                row,
                rows_to_delete,
                f"normalize country={country_cell.value} failed",
            )
        country_cell.value = normalized_country

        phone_str = _phone_to_str(phone_cell.value)
        phone_cell.value = _normalize_phone(phone_str, normalized_country)
        if phone_cell.value == _INVALID_PHONE:
            _add_error_row(
                ws,
                err_ws,
                row,
                rows_to_delete,
                f"parse number={phone_str} country={country_cell.value} failed",
            )
            logger.error(
                f"row: {row} parse number={phone_str} country={country_cell.value} failed"
            )


def _phone_to_str(value) -> str:
    if value is None:
        return _INVALID_PHONE
    if isinstance(value, float):
        return str(int(value))
    return str(value)


def _normalize_country(country) -> str:
    if not country or not isinstance(country, str) or not country.strip():
        return _INVALID_COUNTRY

    country = country.strip()
    country = re.sub(r"\s+", " ", country)

    country = _ALIASES.get(country.lower(), country)

    try:
        res = pycountry.countries.lookup(country)
        return res.alpha_2
    except LookupError:
        return _INVALID_COUNTRY


def _normalize_phone(number: str, country: str) -> str:
    if not number or not country:
        return _INVALID_PHONE

    try:
        parsed = phonenumbers.parse(number, country)
        if not phonenumbers.is_possible_number(parsed):
            return _INVALID_PHONE

        res = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
        return res
    except phonenumbers.NumberParseException:
        return _INVALID_PHONE
