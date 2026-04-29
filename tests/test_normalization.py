import pytest
from datetime import datetime
from openpyxl import Workbook
from openpyxl.cell import Cell
from struct_excel.normalization import (
    _normalize_country,
    _normalize_phone,
    _phone_to_str,
    _trim_cells,
)


class TestNormalizeCountry:
    def test_normalize_country_usa(self):
        result = _normalize_country("USA")
        assert result == "US"

    def test_normalize_country_uk_alias(self):
        result = _normalize_country("UK")
        assert result == "GB"

    def test_normalize_country_gb(self):
        result = _normalize_country("GB")
        assert result == "GB"

    def test_normalize_country_malaysia(self):
        result = _normalize_country("MY")
        assert result == "MY"

    def test_normalize_country_with_whitespace(self):
        result = _normalize_country("  USA  ")
        assert result == "US"

    def test_normalize_country_invalid(self):
        assert _normalize_country("INVALID_COUNTRY") == "UNKNOWN"

    def test_normalize_country_empty(self):
        assert _normalize_country("") == "UNKNOWN"


class TestNormalizePhone:
    def test_normalize_phone_malaysia(self):
        result = _normalize_phone("+60123456789", "MY")
        assert result == "+60123456789"

    def test_normalize_phone_us(self):
        result = _normalize_phone("+1-234-567-8900", "US")
        assert result == "+12345678900"

    def test_normalize_phone_invalid(self):
        assert _normalize_phone("invalid", "US") == "NAN"

    def test_normalize_phone_empty_number(self):
        assert _normalize_phone("", "US") == "NAN"

    def test_normalize_phone_empty_country(self):
        assert _normalize_phone("1234567890", "") == "NAN"


class TestPhoneToStr:
    def test_phone_to_str_none(self):
        result = _phone_to_str(None)
        assert result is "NAN"

    def test_phone_to_str_float(self):
        result = _phone_to_str(1234567890.0)
        assert result == "1234567890"

    def test_phone_to_str_string(self):
        result = _phone_to_str("1234567890")
        assert result == "1234567890"

    def test_phone_to_str_with_plus(self):
        result = _phone_to_str("+60123456789")
        assert result == "+60123456789"


class TestTrimCells:
    def test_trim_cells_removes_zwj(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["test\u200dvalue"])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value == "testvalue"

    def test_trim_cells_removes_zwnj(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["test\u200cvalue"])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value == "testvalue"

    def test_trim_cells_removes_bom(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["\ufefftestvalue"])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value == "testvalue"

    def test_trim_cells_strips_whitespace(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["  testvalue  "])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value == "testvalue"

    def test_trim_cells_ignores_none(self):
        wb = Workbook()
        ws = wb.active
        ws.append([None])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value is None

    def test_trim_cells_ignores_non_string(self):
        wb = Workbook()
        ws = wb.active
        ws.append([123])

        result_ws = _trim_cells(ws)
        cell_value = result_ws.cell(row=1, column=1).value

        assert cell_value == 123
